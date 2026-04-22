import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

BLUE = "2563EB"
LIGHT_BLUE = "EFF4FF"
DARK = "1A1A1A"
GRAY = "6B6B6B"
LIGHT_GRAY = "F7F7F5"
GREEN = "16A34A"
GREEN_BG = "F0FDF4"
YELLOW_BG = "FFF9E6"
WHITE = "FFFFFF"
BORDER_COLOR = "E2E2E0"

title_font = Font(name="Calibri", size=14, bold=True, color=DARK)
section_font = Font(name="Calibri", size=12, bold=True, color=BLUE)
header_font = Font(name="Calibri", size=10, bold=True, color=GRAY)
body_font = Font(name="Calibri", size=11, color=DARK)
hint_font = Font(name="Calibri", size=10, italic=True, color=GRAY)
input_font = Font(name="Calibri", size=11, color=BLUE)

fill_header = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
fill_input = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_green = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
fill_light = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center")


def style_cell(ws, row, col, value, font=body_font, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def add_section_header(ws, row, title):
    style_cell(ws, row, 1, title, font=section_font)
    return row + 1


def add_table_header(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        style_cell(ws, row, col_start + i, h, font=header_font, fill=fill_header, border=thin_border,
                   alignment=center_align)
    return row + 1


def add_input_row(ws, row, label, value, hint="", col_start=1, is_input=True):
    style_cell(ws, row, col_start, label, font=body_font, border=thin_border)
    cell = style_cell(ws, row, col_start + 1, value,
                      font=input_font if is_input else body_font,
                      fill=fill_input if is_input else None,
                      border=thin_border)
    if hint:
        style_cell(ws, row, col_start + 2, hint, font=hint_font)
    return row + 1


def add_tier_row(ws, row, from_m, to_m, method, value, col_start=1, is_input=True):
    fill = fill_input if is_input else None
    font = input_font if is_input else body_font
    style_cell(ws, row, col_start, from_m, font=font, fill=fill, border=thin_border, alignment=center_align)
    style_cell(ws, row, col_start + 1, to_m, font=font, fill=fill, border=thin_border, alignment=center_align)
    style_cell(ws, row, col_start + 2, method, font=font, fill=fill, border=thin_border, alignment=center_align)
    style_cell(ws, row, col_start + 3, value, font=font, fill=fill, border=thin_border, alignment=center_align)
    return row + 1


def make_dropdown(options_list):
    formula = ",".join(options_list)
    dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
    dv.error = "Please select from the dropdown"
    dv.errorTitle = "Invalid entry"
    dv.prompt = "Choose from the list"
    dv.promptTitle = "Select"
    return dv


DV_YES_NO = ["Yes", "No"]
DV_CONTRACT_TYPE = ["Indefinite", "Fixed Term"]
DV_MTA_DEFAULT = ["No", "Yes — always", "Yes — above tenure threshold"]
DV_TIER_METHOD = ["Fixed", "Per year of service"]
DV_VACATION_TYPE = ["All Accrued", "Fixed number of days"]


def add_dropdown(ws, cell_ref, options_list):
    dv = make_dropdown(options_list)
    dv.add(cell_ref)
    ws.add_data_validation(dv)


def build_country_sheet(ws, name, data):
    ws.sheet_properties.tabColor = BLUE
    for c in range(1, 10):
        ws.column_dimensions[get_column_letter(c)].width = 22
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 28

    row = 1
    style_cell(ws, row, 1, f"Termination Liability Configuration — {name}", font=title_font)
    row += 2

    # SECTION 1: General Inputs
    row = add_section_header(ws, row, "1. General Inputs")
    row = add_input_row(ws, row, "Country", data["country"], data.get("country_hint", ""))

    contract_row = row
    row = add_input_row(ws, row, "Contract Type", data["contract_type"],
                        "Indefinite / Fixed Term")
    add_dropdown(ws, ws.cell(row=contract_row, column=2).coordinate, DV_CONTRACT_TYPE)

    if data.get("entity_type"):
        row = add_input_row(ws, row, "Entity Type", data["entity_type"],
                            "Only for countries with multiple employment models")
    else:
        row = add_input_row(ws, row, "Entity Type", "N/A",
                            "Only for DE, FR, BE — leave N/A otherwise", is_input=False)

    mta_row = row
    row = add_input_row(ws, row, "Is MTA the default practice?", data["mta_default"],
                        "No / Yes — always / Yes — above tenure threshold")
    add_dropdown(ws, ws.cell(row=mta_row, column=2).coordinate, DV_MTA_DEFAULT)
    row += 1

    # SECTION 2: Severance Cost Components
    row = add_section_header(ws, row, "2. Severance Cost Components")

    if data["contract_type"] == "Fixed Term":
        remaining_row = row
        row = add_input_row(ws, row, "Use remaining days on contract as severance?",
                            data.get("use_remaining_contract_days", "No"), "Yes / No — if Yes, skip tier table below")
        add_dropdown(ws, ws.cell(row=remaining_row, column=2).coordinate, DV_YES_NO)
        row += 1

    for ci, comp in enumerate(data["components"]):
        style_cell(ws, row, 1,
                   f"Component {ci + 1}: {comp['label']}" + (" (default)" if comp.get("is_default") else ""),
                   font=Font(name="Calibri", size=11, bold=True, color=DARK))
        row += 1
        prorate_row = row
        row = add_input_row(ws, row, "Prorate for partial years?", comp.get("prorate", "Yes"),
                            "Yes (default) / No — e.g., 30 days/yr with 8 months: Yes → 20 days, No → 0 days")
        add_dropdown(ws, ws.cell(row=prorate_row, column=2).coordinate, DV_YES_NO)
        row = add_table_header(ws, row, ["From (tenure months)", "To (tenure months)", "Method", "Value"])
        for t in comp["tiers"]:
            method_row = row
            row = add_tier_row(ws, row, t[0], t[1], t[2], t[3])
            add_dropdown(ws, ws.cell(row=method_row, column=3).coordinate, DV_TIER_METHOD)
        style_cell(ws, row, 1, "Max cap (days)", font=body_font, border=thin_border)
        style_cell(ws, row, 2, comp.get("max_cap", "None"), font=input_font, fill=fill_input, border=thin_border)
        row += 1
        style_cell(ws, row, 1, "↳ Add more tiers by inserting rows above this line", font=hint_font)
        row += 2

    style_cell(ws, row, 1, "↳ To add another component, copy a component block above and change the label",
               font=hint_font)
    row += 2

    # SECTION 3: MTA Configuration (if applicable)
    row = add_section_header(ws, row, "3. MTA Configuration (if MTA = Yes)")
    if data["mta_default"] in ("Yes — always", "Yes — above tenure threshold"):
        warn_font = Font(name="Calibri", size=10, italic=True, color="9A3412")
        warn_fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
        style_cell(ws, row, 1, "⚠ When MTA is used, these severance days REPLACE the cost components above — they are not added together.",
                   font=warn_font, fill=warn_fill)
        row += 1
        if data["mta_default"] == "Yes — above tenure threshold":
            row = add_input_row(ws, row, "MTA applies when tenure ≥", data.get("mta_tenure_threshold", ""),
                                "months")
        row = add_input_row(ws, row, "Days to secure MTA (excl. notice period)", data["mta_days"],
                            "Fixed number of days")
        row += 1
        style_cell(ws, row, 1, "MTA Severance Tiers", font=Font(name="Calibri", size=11, bold=True, color=DARK))
        row += 1
        tier_headers = ["From (tenure months)", "To (tenure months)", "Method", "Value"]
        row = add_table_header(ws, row, tier_headers)
        for t in data["mta_tiers"]:
            method_row = row
            row = add_tier_row(ws, row, t[0], t[1], t[2], t[3])
            add_dropdown(ws, ws.cell(row=method_row, column=3).coordinate, DV_TIER_METHOD)
        style_cell(ws, row, 1, "Max cap (days)", font=body_font, border=thin_border)
        style_cell(ws, row, 2, data.get("mta_max_cap", "None"), font=input_font, fill=fill_input, border=thin_border)
        row += 1
        style_cell(ws, row, 1, "↳ Add more tiers as needed by inserting rows above", font=hint_font)
    else:
        style_cell(ws, row, 1, "N/A — MTA is not the default for this country", font=hint_font)
    row += 2

    # SECTION 4: Vacation Payout
    row = add_section_header(ws, row, "4. Vacation Payout")
    style_cell(ws, row, 1,
               "For deposit sizing, we use the statutory entitlement from the EA/config — not the T&A system balance (which HR does not trust).",
               font=hint_font)
    row += 1
    vac_enabled_row = row
    row = add_input_row(ws, row, "Pay out accrued vacation during termination?",
                        data["vacation"]["enabled"], "Yes / No")
    add_dropdown(ws, ws.cell(row=vac_enabled_row, column=2).coordinate, DV_YES_NO)
    if data["vacation"]["enabled"] == "Yes":
        vac_type_row = row
        row = add_input_row(ws, row, "Statutory minimum payout type",
                            data["vacation"]["minimum"], "All Accrued / Fixed number of days")
        add_dropdown(ws, ws.cell(row=vac_type_row, column=2).coordinate, DV_VACATION_TYPE)
        if data["vacation"]["minimum"] == "Fixed number of days":
            row += 1
            style_cell(ws, row, 1, "Vacation Payout Tiers (statutory entitlement)",
                       font=Font(name="Calibri", size=11, bold=True, color=DARK))
            row += 1
            row = add_table_header(ws, row, ["From (tenure months)", "To (tenure months)", "Method", "Value"])
            for t in data["vacation"]["tiers"]:
                method_row = row
                row = add_tier_row(ws, row, t[0], t[1], t[2], t[3])
                add_dropdown(ws, ws.cell(row=method_row, column=3).coordinate, DV_TIER_METHOD)
            style_cell(ws, row, 1, "Max cap (days)", font=body_font, border=thin_border)
            style_cell(ws, row, 2, data["vacation"].get("max_cap", "None"),
                       font=input_font, fill=fill_input, border=thin_border)
            row += 1
            style_cell(ws, row, 1, "↳ Add more tiers by inserting rows above this line", font=hint_font)
        elif data["vacation"]["minimum"] == "All Accrued":
            row += 1
            style_cell(ws, row, 1,
                       "For deposit sizing: we assume worst case = full annual entitlement from the EA (prorated to tenure). Actual usage is unknown at estimation time.",
                       font=hint_font)
    row += 2

    # SECTION 5: Compensation Variable Metadata
    row = add_section_header(ws, row, "5. Compensation Variable Metadata")
    style_cell(ws, row, 1,
               "For each compensation variable, indicate Yes/No for each column. New variables default to No.",
               font=hint_font)
    row += 1
    cv_headers = ["Compensation Variable", "Included in Working Notice?", "Included in Severance?",
                  "Included in Vacation Pay?"]
    row = add_table_header(ws, row, cv_headers)
    for cv in data["comp_vars"]:
        style_cell(ws, row, 1, cv[0], font=body_font, border=thin_border)
        for j in range(1, 4):
            style_cell(ws, row, j + 1, cv[j], font=input_font, fill=fill_input, border=thin_border,
                       alignment=center_align)
            add_dropdown(ws, ws.cell(row=row, column=j + 1).coordinate, DV_YES_NO)
        row += 1
    style_cell(ws, row, 1, "↳ Add more variables by inserting rows above", font=hint_font)
    row += 2

    # SECTION 6: Worked Examples
    examples = data.get("examples", [])
    if examples:
        row = add_section_header(ws, row, "6. Worked Examples — How Severance & Vacation Pay Are Calculated")
        style_cell(ws, row, 1,
                   "These examples show how the config above translates into actual severance days and dollar amounts.",
                   font=hint_font)
        row += 2

        example_label_font = Font(name="Calibri", size=11, bold=True, color=DARK)
        step_font = Font(name="Calibri", size=11, color=DARK)
        result_font = Font(name="Calibri", size=11, bold=True, color=GREEN)
        calc_font = Font(name="Calibri", size=11, color=BLUE)

        for ex_idx, ex in enumerate(examples):
            style_cell(ws, row, 1, f"Example {ex_idx + 1}: {ex['title']}", font=section_font)
            row += 1

            # Employee profile
            style_cell(ws, row, 1, "Employee Profile", font=example_label_font)
            row += 1
            for label, val in ex["profile"]:
                style_cell(ws, row, 1, f"  {label}", font=step_font, border=thin_border)
                style_cell(ws, row, 2, val, font=calc_font, border=thin_border)
                row += 1
            row += 1

            # Severance calculation
            style_cell(ws, row, 1, "Severance Calculation", font=example_label_font)
            row += 1
            for step in ex["severance_steps"]:
                if step.startswith("→"):
                    style_cell(ws, row, 1, step, font=result_font)
                elif step.startswith("="):
                    style_cell(ws, row, 1, f"  {step}", font=calc_font)
                else:
                    style_cell(ws, row, 1, f"  {step}", font=step_font)
                row += 1
            row += 1

            # Vacation calculation
            if ex.get("vacation_steps"):
                style_cell(ws, row, 1, "Vacation Payout Calculation", font=example_label_font)
                row += 1
                for step in ex["vacation_steps"]:
                    if step.startswith("→"):
                        style_cell(ws, row, 1, step, font=result_font)
                    elif step.startswith("="):
                        style_cell(ws, row, 1, f"  {step}", font=calc_font)
                    else:
                        style_cell(ws, row, 1, f"  {step}", font=step_font)
                    row += 1
                row += 1

            # Total
            if ex.get("total_steps"):
                style_cell(ws, row, 1, "What Gets Passed to Risk", font=example_label_font)
                row += 1
                for step in ex["total_steps"]:
                    if step.startswith("→"):
                        style_cell(ws, row, 1, step, font=result_font)
                    else:
                        style_cell(ws, row, 1, f"  {step}", font=step_font)
                    row += 1
            row += 2

        style_cell(ws, row, 1,
                   "Note: Risk converts severance_days into a dollar amount using the severance salary basis (comp variables marked 'Included in Severance = Yes').",
                   font=hint_font)
        row += 1
        style_cell(ws, row, 1,
                   "Note: Risk converts vacation_days_payout into a dollar amount using the vacation pay salary basis (comp variables marked 'Included in Vacation Pay = Yes').",
                   font=hint_font)
        row += 2

    # SECTION 7: Notes / Items Not Captured
    section_num = 7 if examples else 6
    row = add_section_header(ws, row, f"{section_num}. Notes / Items Not Captured")
    style_cell(ws, row, 1, "Use this section to flag anything the config above cannot express:", font=hint_font)
    row += 1
    for note in data.get("notes", []):
        style_cell(ws, row, 1, f"• {note}", font=body_font)
        row += 1

    return ws


# ========== INSTRUCTIONS TAB ==========
ws_instr = wb.active
ws_instr.title = "Instructions"
ws_instr.sheet_properties.tabColor = "16A34A"
ws_instr.column_dimensions["A"].width = 80

call_to_action_font = Font(name="Calibri", size=11, bold=True, color="DC2626")
step_bold_font = Font(name="Calibri", size=11, bold=True, color=DARK)

instructions = [
    ("Termination Liability Configuration", title_font),
    ("Data collection sheet — to be filled by HR Country Managers", Font(name="Calibri", size=12, italic=True, color=GRAY)),
    ("", body_font),

    ("What is this and why are we doing it?", section_font),
    ("", body_font),
    ("When we onboard an EOR employee, we collect a deposit from the client to cover potential termination", body_font),
    ("costs. Today, that deposit amount is often inaccurate — it may not reflect the actual termination", body_font),
    ("liability, leaving Rippling exposed to financial risk if the employee is terminated.", body_font),
    ("", body_font),
    ("The root cause is that termination rules vary by country, and we have no structured way to capture", body_font),
    ("them. Severance formulas, vacation payout obligations, and MTA practices all differ by country and", body_font),
    ("change based on how long an employee has been working.", body_font),
    ("", body_font),
    ("We are building a system that takes these rules as structured configuration and automatically", body_font),
    ("calculates the correct deposit amount per employee. To do that, we first need to collect the rules", body_font),
    ("from you. Your input will directly feed the product.", body_font),
    ("", body_font),
    ("UI preview (read-only):", Font(name="Calibri", size=11, bold=True, color=DARK)),
    ("See how this will look once it's live in Control Center:", body_font),
    ("https://sanjkris.github.io/termination-liability-config/", Font(name="Calibri", size=11, underline="single", color=BLUE)),
    ("", body_font),

    ("What we need from you", section_font),
    ("For each country you manage, please fill out one tab with the termination rules.", call_to_action_font),
    ("", body_font),
    ("We understand that terminations are typically handled case by case. The goal here is not to capture", body_font),
    ("every edge case — it is to get a general sense of the rules we typically follow, so we can size", body_font),
    ("deposits more accurately.", body_font),
    ("", body_font),
    ("We need four things for each country. Each is explained below.", body_font),
    ("", body_font),

    ("", body_font),
    ("1. Severance Pay", section_font),
    ("When we terminate an employee, how many days of pay do they receive as severance?", body_font),
    ("", body_font),
    ("Questions to answer:", Font(name="Calibri", size=11, bold=True, color=DARK)),
    ("  • Does the employee receive any severance at all? If not, enter 0.", body_font),
    ("  • Does the amount change based on how long they have been employed?", body_font),
    ("    (e.g., 0 days if under 6 months, 30 days per year of service after 6 months)", body_font),
    ("  • Is there a maximum cap? (e.g., Spain caps objective dismissal severance at 360 days)", body_font),
    ("  • Are there multiple severance components that get added together?", body_font),
    ("", body_font),
    ("Example — single component:", Font(name="Calibri", size=11, italic=True, color=DARK)),
    ("  Philippines: Redundancy pay = 30 days per year of service (no cap).", body_font),
    ("  Employees under 6 months get nothing.", body_font),
    ("  → One component ('Redundancy Pay') with two tiers: 0–6 months = 0 days, 6+ months = 30 days/yr.", body_font),
    ("", body_font),
    ("Example — multiple components:", Font(name="Calibri", size=11, italic=True, color=DARK)),
    ("  Spain: Objective dismissal = 20 days/yr (capped at 360 days) + pagas extra = 2.5 days/yr (no cap).", body_font),
    ("  → Two separate components, each with its own tier table. They are added together.", body_font),
    ("", body_font),
    ("If your country has more than one type of payment on termination (redundancy pay, gratuity premium,", body_font),
    ("back pay, etc.), add each one as a separate component in Section 2 of your country tab.", body_font),
    ("", body_font),
    ("Fixed-term contracts:", Font(name="Calibri", size=11, bold=True, color=DARK)),
    ("  If the contract type is 'Fixed Term', Section 2 will show an additional toggle:", body_font),
    ("  'Use remaining days on contract as severance?' — Yes / No", body_font),
    ("  If Yes, severance = remaining days left on the contract (contract end date − termination date).", body_font),
    ("  In that case, you can skip the tier table — the system calculates it automatically.", body_font),
    ("  If No, fill in the tier table as you would for an indefinite contract.", body_font),
    ("", body_font),
    ("  Rules may differ between indefinite and fixed-term contracts in the same country.", body_font),
    ("  If so, create a separate tab for each — see the tab naming convention below.", body_font),
    ("", body_font),

    ("", body_font),
    ("2. MTA (Mutual Termination Agreement)", section_font),
    ("An MTA is when Rippling negotiates a mutual separation with the employee instead of a unilateral", body_font),
    ("termination. It typically involves a negotiated severance package.", body_font),
    ("", body_font),
    ("Questions to answer:", Font(name="Calibri", size=11, bold=True, color=DARK)),
    ("  • Does Rippling typically negotiate an MTA in this country?", body_font),
    ("     - 'No' — we do not use MTAs here", body_font),
    ("     - 'Yes — always' — we always negotiate an MTA regardless of tenure", body_font),
    ("     - 'Yes — above tenure threshold' — we only use MTA for employees above a certain tenure", body_font),
    ("       (e.g., above 2 years). Please specify the threshold in months.", body_font),
    ("  • How many days does it typically take to secure the MTA? (excluding applicable notice period days)", body_font),
    ("    This is the negotiation time — from when we start the process to when it is signed.", body_font),
    ("  • What is the typical MTA severance? Express it using the tier model (explained below).", body_font),
    ("", body_font),
    ("Key point — MTA vs. standard severance:", Font(name="Calibri", size=11, bold=True, color="DC2626")),
    ("  These are mutually exclusive. For any given employee, only one path applies:", body_font),
    ("  • If MTA is used → the MTA severance applies (Section 3 of the tab)", body_font),
    ("  • If MTA is not used → the standard severance cost components apply (Section 2)", body_font),
    ("  They are never added together. The system applies one or the other.", body_font),
    ("", body_font),
    ("Example:", Font(name="Calibri", size=11, italic=True, color=DARK)),
    ("  UK: MTA is used only for employees above 24 months of tenure. Below 24 months, standard", body_font),
    ("  statutory redundancy applies (which is 0 for under 2 years). MTA severance is typically", body_font),
    ("  30–90 days. Negotiation takes around 10 working days.", body_font),
    ("", body_font),

    ("", body_font),
    ("3. Vacation Payout", section_font),
    ("When an employee is terminated, are they entitled to a payout for unused vacation days?", body_font),
    ("", body_font),
    ("Questions to answer:", Font(name="Calibri", size=11, bold=True, color=DARK)),
    ("  • Is vacation paid out on termination at all?", body_font),
    ("  • If yes: do we pay the full accrued (unused) balance, or is there a statutory minimum?", body_font),
    ("    What do we typically advise clients to budget for?", body_font),
    ("  • If there is a fixed minimum (not the full balance), express it using the tier model.", body_font),
    ("", body_font),
    ("Example — full accrued balance:", Font(name="Calibri", size=11, italic=True, color=DARK)),
    ("  Philippines: All accrued unused vacation is paid out on termination.", body_font),
    ("  → Select 'All Accrued' in Section 4.", body_font),
    ("", body_font),
    ("Example — fixed statutory minimum:", Font(name="Calibri", size=11, italic=True, color=DARK)),
    ("  A country where employees are entitled to at least 5 days of vacation payout regardless of balance.", body_font),
    ("  → Select 'Fixed number of days' and fill in the tier table with the minimum.", body_font),
    ("", body_font),

    ("", body_font),
    ("4. Compensation Variables", section_font),
    ("Not all parts of an employee's compensation are included in every calculation. For example, base", body_font),
    ("salary is typically included in severance, but a signing bonus probably is not.", body_font),
    ("", body_font),
    ("For each compensation variable (base salary, target bonus, commission, allowances, etc.), tell us:", body_font),
    ("", body_font),
    ("  • Included in Working Notice? — Is this variable paid during the notice period?", body_font),
    ("    Example: In most countries, the employee receives full base salary and allowances during notice.", body_font),
    ("", body_font),
    ("  • Included in Severance? — Is this variable part of the daily rate used to calculate severance pay?", body_font),
    ("    Example: In the Philippines, severance is based on base salary only. Allowances are excluded.", body_font),
    ("", body_font),
    ("  • Included in Vacation Pay? — Is this variable part of the daily rate used to calculate vacation payout?", body_font),
    ("    Example: If vacation payout is based on base salary only, mark only base salary as 'Yes'.", body_font),
    ("", body_font),
    ("Fill in the table in Section 5 of your country tab. List every compensation variable that applies and", body_font),
    ("mark Yes or No for each column. If a variable does not exist in your country, skip it.", body_font),
    ("", body_font),

    ("", body_font),
    ("How to fill this out", section_font),
    ("Step 1:  Go to the 'Template' tab at the bottom of this workbook.", step_bold_font),
    ("Step 2:  Right-click the tab → Duplicate → rename it using the naming convention below.", step_bold_font),
    ("Step 3:  Fill in the yellow-highlighted cells.", step_bold_font),
    ("         These are the cells that need your input. Everything else is pre-filled or instructional.", body_font),
    ("Step 4:  Add or remove tier rows as needed.", step_bold_font),
    ("         Insert rows within the tier tables if your country has more brackets.", body_font),
    ("Step 5:  If there are additional severance components beyond redundancy pay, add them.", step_bold_font),
    ("         For example, if the country also requires a gratuity premium or 13th-month back pay,", body_font),
    ("         copy a component block and change the label. Each component has its own tier table.", body_font),
    ("Step 6:  Use the 'Notes' section at the bottom for anything the template cannot express.", step_bold_font),
    ("         Edge cases, caveats, probation rules, special law references — put them here.", body_font),
    ("", body_font),

    ("Tab naming convention", section_font),
    ("Name each tab using this pattern:", body_font),
    ("", body_font),
    ("  CountryCode_EntityType_Province_ContractType", Font(name="Courier New", size=11, bold=True, color=BLUE)),
    ("", body_font),
    ("Omit fields that don't apply — but always include the country code and contract type.", body_font),
    ("", body_font),
    ("Examples:", Font(name="Calibri", size=11, italic=True, color=DARK)),
    ("  • PH_Indefinite                — Philippines, no entity/province variation", body_font),
    ("  • ES_Indefinite                — Spain, indefinite contracts", body_font),
    ("  • DE_AUG_Indefinite            — Germany, AUG entity type", body_font),
    ("  • DE_Consultant_Fixed Term     — Germany, Consultant entity, fixed-term contracts", body_font),
    ("  • CA_Ontario_Indefinite        — Canada, Ontario province", body_font),
    ("  • CA_Alberta_Indefinite        — Canada, Alberta province", body_font),
    ("", body_font),
    ("If termination rules differ by entity type, province, or contract type within the same country,", body_font),
    ("create a separate tab for each variation. This ensures we configure the correct rules for each combination.", body_font),
    ("", body_font),

    ("Pre-filled reference examples", section_font),
    ("These tabs show what a completed country looks like. Use them as a guide:", body_font),
    ("", body_font),
    ("• PH_Indefinite — simple case: no MTA, per-year severance with a 6-month threshold, all accrued vacation", body_font),
    ("• ES_Indefinite — cap example: objective dismissal at 20 days/yr capped at 360 days", body_font),
    ("", body_font),
    ("Each reference tab includes a 'Worked Examples' section at the bottom. These show step by step how", body_font),
    ("the rules translate into actual severance days and amounts for sample employees.", body_font),
    ("", body_font),

    ("How the tier model works", section_font),
    ("Throughout this spreadsheet you will see 'tier tables.' These are how we express rules that change", body_font),
    ("based on how long the employee has been working (their tenure).", body_font),
    ("", body_font),
    ("Each tier has four columns:", Font(name="Calibri", size=11, bold=True, color=DARK)),
    ("", body_font),
    ("  From (tenure months)  — when this bracket starts", body_font),
    ("  To (tenure months)    — when this bracket ends (use 'No limit' if it goes on indefinitely)", body_font),
    ("  Method                — how the payout is calculated in this bracket:", body_font),
    ("     'Fixed'               = a flat number of days, regardless of exact tenure", body_font),
    ("     'Per year of service' = days multiplied by years of tenure (e.g., 30 days × 2.5 years = 75 days)", body_font),
    ("  Value                 — the number of days", body_font),
    ("", body_font),
    ("  Prorate for partial years? — Yes (default) or No.", body_font),
    ("  Yes: partial tenure is prorated (e.g., 30 days/yr with 8 months tenure → 30 × 8/12 = 20 days).", body_font),
    ("  No: only completed years count (e.g., 30 days/yr with 8 months tenure → 0 completed years = 0 days).", body_font),
    ("", body_font),
    ("  Max cap (days) — appears below the tier table. This is the overall ceiling for the component.", body_font),
    ("  If there is no cap, write 'None'.", body_font),
    ("", body_font),

    ("Example 1 — Simple (Philippines redundancy pay):", Font(name="Calibri", size=11, bold=True, color=DARK)),
    ("", body_font),
    ("  Rule: No severance for the first 6 months. After 6 months, 30 days per year of service.", body_font),
    ("", body_font),
    ("  How to fill the tier table:", body_font),
    ("  ┌──────────────────┬──────────────────┬─────────────────────┬──────────────────────┐", Font(name="Courier New", size=10, color=DARK)),
    ("  │ From (months)    │ To (months)      │ Method              │ Value                │", Font(name="Courier New", size=10, color=DARK)),
    ("  ├──────────────────┼──────────────────┼─────────────────────┼──────────────────────┤", Font(name="Courier New", size=10, color=DARK)),
    ("  │ 0                │ 6                │ Fixed               │ 0 days               │", Font(name="Courier New", size=10, color=DARK)),
    ("  │ 6                │ No limit         │ Per year of service │ 30 days              │", Font(name="Courier New", size=10, color=DARK)),
    ("  └──────────────────┴──────────────────┴─────────────────────┴──────────────────────┘", Font(name="Courier New", size=10, color=DARK)),
    ("  Max cap: None", body_font),
    ("", body_font),
    ("  What this means for actual employees:", body_font),
    ("    3 months tenure  → tier 1 (0–6) → Fixed 0 days → no severance", body_font),
    ("    8 months tenure  → tier 2 (6+)  → 30 × (8/12) = 20 days", body_font),
    ("    2.5 years tenure → tier 2 (6+)  → 30 × 2.5 = 75 days", body_font),
    ("", body_font),

    ("Example 2 — With a cap (Spain objective dismissal):", Font(name="Calibri", size=11, bold=True, color=DARK)),
    ("", body_font),
    ("  Rule: 20 days per year of service, capped at 360 days.", body_font),
    ("", body_font),
    ("  ┌──────────────────┬──────────────────┬─────────────────────┬──────────────────────┐", Font(name="Courier New", size=10, color=DARK)),
    ("  │ From (months)    │ To (months)      │ Method              │ Value                │", Font(name="Courier New", size=10, color=DARK)),
    ("  ├──────────────────┼──────────────────┼─────────────────────┼──────────────────────┤", Font(name="Courier New", size=10, color=DARK)),
    ("  │ 0                │ No limit         │ Per year of service │ 20 days              │", Font(name="Courier New", size=10, color=DARK)),
    ("  └──────────────────┴──────────────────┴─────────────────────┴──────────────────────┘", Font(name="Courier New", size=10, color=DARK)),
    ("  Max cap: 360 days", body_font),
    ("", body_font),
    ("  What this means:", body_font),
    ("    5 years tenure  → 20 × 5 = 100 days (under the cap)", body_font),
    ("    20 years tenure → 20 × 20 = 400 days → capped at 360 days", body_font),
    ("", body_font),

    ("Example 3 — Proration (if your country prorates within a year):", Font(name="Calibri", size=11, bold=True, color=DARK)),
    ("", body_font),
    ("  If proration applies, please break it out into explicit tiers. For example, if severance is", body_font),
    ("  30 days per year of service prorated monthly:", body_font),
    ("", body_font),
    ("  ┌──────────────────┬──────────────────┬─────────────────────┬──────────────────────┐", Font(name="Courier New", size=10, color=DARK)),
    ("  │ From (months)    │ To (months)      │ Method              │ Value                │", Font(name="Courier New", size=10, color=DARK)),
    ("  ├──────────────────┼──────────────────┼─────────────────────┼──────────────────────┤", Font(name="Courier New", size=10, color=DARK)),
    ("  │ 0                │ 1                │ Fixed               │ 2.5 days             │", Font(name="Courier New", size=10, color=DARK)),
    ("  │ 1                │ 2                │ Fixed               │ 5 days               │", Font(name="Courier New", size=10, color=DARK)),
    ("  │ 2                │ 6                │ Fixed               │ 15 days              │", Font(name="Courier New", size=10, color=DARK)),
    ("  │ 6                │ 12               │ Fixed               │ 30 days              │", Font(name="Courier New", size=10, color=DARK)),
    ("  │ 12               │ No limit         │ Per year of service │ 30 days              │", Font(name="Courier New", size=10, color=DARK)),
    ("  └──────────────────┴──────────────────┴─────────────────────┴──────────────────────┘", Font(name="Courier New", size=10, color=DARK)),
    ("", body_font),
    ("  Use as many or as few tiers as needed to represent how your country actually works.", body_font),
    ("", body_font),

    ("Rules for tier tables:", Font(name="Calibri", size=11, bold=True, color=DARK)),
    ("  • Tiers must be contiguous — no gaps (e.g., 0–12 then 24–60 is invalid; 12–24 must exist).", body_font),
    ("  • Fractional days are fine (e.g., 2.5 days, 1.5 days per year of service).", body_font),
    ("  • Add as many rows as you need — just insert rows within the table.", body_font),
    ("", body_font),

    ("What we are not asking for (out of scope for v1)", section_font),
    ("• Historical salary lookbacks or trailing averages — we use current compensation only", body_font),
    ("• Financial reserve payments (e.g., 10% of salary set aside monthly)", body_font),
    ("• Country-specific special pay types (e.g., Colombia integral salary)", body_font),
    ("If any of these are critical for your country, please note it in the Notes section.", body_font),
    ("", body_font),

    ("Questions or something that does not fit?", section_font),
    ("Slack Sanjana Krishnam, or add a comment in the Notes section of your country tab.", body_font),
    ("If a rule cannot be expressed in the template, describe it in Notes — that is exactly what we need to know.", body_font),
]

for i, (text, font) in enumerate(instructions):
    cell = ws_instr.cell(row=i + 1, column=1, value=text)
    cell.font = font


# ========== PHILIPPINES TAB ==========
ph_data = {
    "country": "Philippines (PH)",
    "contract_type": "Indefinite",
    "entity_type": None,
    "mta_default": "No",  
    "mta_days": None,
    "mta_tiers": [],
    "components": [
        {
            "label": "Redundancy Pay",
            "is_default": True,
            "tiers": [
                (0, 6, "Fixed", "0 days"),
                (6, "No limit", "Per year of service", "30 days per year of service"),
            ],
            "max_cap": "None",
        }
    ],
    "vacation": {
        "enabled": "Yes",
        "minimum": "All Accrued",
        "tiers": [],
    },
    "comp_vars": [
        ("Gross Base Salary", "Yes", "Yes", "Yes"),
        ("Target Bonus", "Yes", "No", "No"),
        ("Signing Bonus", "Yes", "No", "No"),
        ("On-Target Commission", "Yes", "No", "No"),
        ("Rice Allowance", "Yes", "No", "No"),
        ("Uniform & Clothing Allowance", "Yes", "No", "No"),
        ("Laundry Allowance", "Yes", "No", "No"),
        ("Medical Cash Allowance", "Yes", "No", "No"),
        ("Internet Allowance", "Yes", "No", "No"),
        ("Productivity Incentive Allowance", "Yes", "No", "No"),
        ("Christmas / Anniversary Gift", "Yes", "No", "No"),
        ("13th Month Salary", "Yes", "No", "No"),
    ],
    "examples": [
        {
            "title": "Employee with 3 months tenure (below 6-month threshold — no severance)",
            "profile": [
                ("Tenure", "3 months"),
                ("Gross Base Salary", "₱80,000 / month"),
                ("Accrued Vacation Balance", "3 days"),
            ],
            "severance_steps": [
                "Step 1: Identify the matching tier for tenure = 3 months",
                "  Tier 1: 0 → 6 months, Method: Fixed, Value: 0 days",
                "  Employee is within the first 6 months → no severance entitlement",
                "→ severance_days = 0",
            ],
            "vacation_steps": [
                "Config says: All Accrued",
                "Employee has 3 days accrued",
                "→ vacation_days_payout = 3",
            ],
            "total_steps": [
                "severance_days = 0 (below 6-month threshold)",
                "vacation_days_payout = 3",
                "→ Passed to Risk for deposit calculation",
            ],
        },
        {
            "title": "Employee with 8 months tenure (above threshold — prorated severance)",
            "profile": [
                ("Tenure", "8 months (0.67 years)"),
                ("Gross Base Salary", "₱80,000 / month"),
                ("Accrued Vacation Balance", "5 days"),
            ],
            "severance_steps": [
                "Step 1: Identify the matching tier for tenure = 8 months",
                "  Tier 2: 6 → No limit months, Method: Per year of service, Value: 30 days/yr",
                "Step 2: Calculate severance days (prorated)",
                "  Tenure = 8 months = 8/12 years = 0.67 years",
                "  30 days/yr × 0.67 years = 20 days",
                "Step 3: Check max cap → None",
                "→ severance_days = 20",
                "",
                "Proration breakdown for reference:",
                "  At 6 months: 30 × (6/12) = 15 days",
                "  At 7 months: 30 × (7/12) = 17.5 days",
                "  At 8 months: 30 × (8/12) = 20 days",
                "  At 12 months: 30 × (12/12) = 30 days",
            ],
            "vacation_steps": [
                "Config says: All Accrued",
                "Employee has 5 days accrued",
                "→ vacation_days_payout = 5",
            ],
            "total_steps": [
                "severance_days = 20 (prorated for 8 months)",
                "vacation_days_payout = 5",
                "→ Passed to Risk for deposit calculation",
            ],
        },
        {
            "title": "Employee with 2.5 years tenure (full proration)",
            "profile": [
                ("Tenure", "2 years, 6 months (2.5 years)"),
                ("Gross Base Salary", "₱100,000 / month"),
                ("Accrued Vacation Balance", "18 days"),
            ],
            "severance_steps": [
                "Step 1: Identify the matching tier for tenure = 30 months",
                "  Tier 2: 6 → No limit months, Method: Per year of service, Value: 30 days/yr",
                "Step 2: Calculate severance days (prorated)",
                "  Tenure = 2.5 years × 30 days/yr = 75 days",
                "Step 3: Check max cap → None",
                "→ severance_days = 75",
                "",
                "Step 4: Determine severance salary basis (comp vars where Severance = Yes)",
                "  Only 'Gross Base Salary' is included in severance",
                "  Severance daily rate = ₱100,000 / 30 = ₱3,333.33",
                "Step 5: Calculate severance pay",
                "= 75 days × ₱3,333.33 = ₱250,000",
                "→ Severance pay = ₱250,000",
            ],
            "vacation_steps": [
                "Config says: All Accrued",
                "Employee has 18 days accrued",
                "→ vacation_days_payout = 18",
                "",
                "Vacation pay salary basis (comp vars where Vacation Pay = Yes):",
                "  Only 'Gross Base Salary' is included",
                "  Vacation daily rate = ₱100,000 / 30 = ₱3,333.33",
                "= 18 days × ₱3,333.33 = ₱60,000",
                "→ Vacation payout = ₱60,000",
            ],
            "total_steps": [
                "severance_days = 75 (passed to Risk)",
                "vacation_days_payout = 18 (passed to Risk)",
                "→ Risk uses the salary bases above to convert days → dollar amounts for deposit sizing",
            ],
        },
    ],
    "notes": [
        "0–6 months tenure: no severance entitlement.",
        "6 months and above: 30 days per year of service, prorated for partial years.",
        "In PH, full compensation is paid during notice period, but severance uses base salary only.",
        "13th Month Salary is mandatory and always included in working notice.",
    ],
}

ws_ph = wb.create_sheet("PH_Indefinite")
build_country_sheet(ws_ph, "Philippines", ph_data)


# ========== SPAIN TAB ==========
es_data = {
    "country": "Spain (ES)",
    "contract_type": "Indefinite",
    "entity_type": None,
    "mta_default": "No",
    "mta_days": None,
    "mta_tiers": [],
    "components": [
        {
            "label": "Severance — Objective Dismissal (Procedente)",
            "is_default": True,
            "tiers": [
                (0, "No limit", "Per year of service", "20 days per year of service"),
            ],
            "max_cap": "360",
        },
    ],
    "vacation": {
        "enabled": "Yes",
        "minimum": "All Accrued",
        "tiers": [],
    },
    "comp_vars": [
        ("Gross Base Salary", "Yes", "Yes", "Yes"),
        ("Pagas Extra (13th/14th Month)", "Yes", "Yes", "Yes"),
        ("Target Bonus", "Yes", "No", "No"),
        ("On-Target Commission", "Yes", "No", "No"),
        ("Meal Allowance", "Yes", "No", "No"),
        ("Transport Allowance", "Yes", "No", "No"),
    ],
    "examples": [
        {
            "title": "Employee with 5 years tenure — Objective Dismissal",
            "profile": [
                ("Tenure", "5 years (60 months)"),
                ("Gross Base Salary", "€4,000 / month"),
                ("Pagas Extra (13th/14th)", "€8,000 / year (= €666.67 / month equivalent)"),
                ("Target Bonus", "€6,000 / year"),
                ("Accrued Vacation Balance", "12 days"),
            ],
            "severance_steps": [
                "Step 1: Identify the component → Objective Dismissal",
                "  Tier: 0 → No limit months, Method: Per year of service, Value: 20 days/yr",
                "",
                "Step 2: Calculate severance days",
                "  Tenure = 5 years × 20 days/yr = 100 days",
                "",
                "Step 3: Check max cap",
                "  Max cap = 360 days → 100 < 360, no cap applies",
                "→ severance_days = 100",
                "",
                "Step 4: Determine severance salary basis (comp vars where Severance = Yes)",
                "  Gross Base Salary: €4,000/month ✓",
                "  Pagas Extra: €666.67/month ✓",
                "  Target Bonus: ✗ (not in severance basis)",
                "  Total monthly severance basis = €4,666.67",
                "  Daily rate = €4,666.67 / 30 = €155.56",
                "",
                "Step 5: Calculate severance pay",
                "= 100 days × €155.56 = €15,556",
                "→ Severance pay = €15,556",
            ],
            "vacation_steps": [
                "Config says: All Accrued",
                "Employee has 12 days accrued",
                "→ vacation_days_payout = 12",
                "",
                "Vacation pay salary basis (comp vars where Vacation Pay = Yes):",
                "  Gross Base Salary: €4,000/month ✓",
                "  Pagas Extra: €666.67/month ✓",
                "  Daily rate = €4,666.67 / 30 = €155.56",
                "= 12 days × €155.56 = €1,866.67",
                "→ Vacation payout = €1,866.67",
            ],
            "total_steps": [
                "severance_days = 100 (passed to Risk)",
                "vacation_days_payout = 12 (passed to Risk)",
                "→ Risk uses the salary bases to convert → total deposit exposure = €15,556 + €1,866.67 = €17,423",
            ],
        },
        {
            "title": "Employee with 20 years tenure — Cap applies",
            "profile": [
                ("Tenure", "20 years (240 months)"),
                ("Gross Base Salary", "€5,000 / month"),
                ("Pagas Extra (13th/14th)", "€10,000 / year (= €833.33 / month equivalent)"),
                ("Accrued Vacation Balance", "22 days"),
            ],
            "severance_steps": [
                "Step 1: Identify the component → Objective Dismissal",
                "  Tier: 0 → No limit months, Method: Per year of service, Value: 20 days/yr",
                "",
                "Step 2: Calculate severance days (before cap)",
                "  Tenure = 20 years × 20 days/yr = 400 days",
                "",
                "Step 3: Check max cap",
                "  Max cap = 360 days → 400 > 360, CAP APPLIES",
                "→ severance_days = 360 (capped)",
                "",
                "Step 4: Severance salary basis",
                "  Base + Pagas Extra = €5,000 + €833.33 = €5,833.33/month",
                "  Daily rate = €5,833.33 / 30 = €194.44",
                "",
                "Step 5: Calculate severance pay",
                "= 360 days × €194.44 = €70,000",
                "→ Severance pay = €70,000",
                "",
                "Without the cap it would have been 400 × €194.44 = €77,778 — the cap saves €7,778",
            ],
            "vacation_steps": [
                "Config says: All Accrued",
                "Employee has 22 days accrued",
                "→ vacation_days_payout = 22",
                "= 22 × €194.44 = €4,278",
                "→ Vacation payout = €4,278",
            ],
            "total_steps": [
                "severance_days = 360 (capped from 400)",
                "vacation_days_payout = 22",
                "→ Total deposit exposure = €70,000 + €4,278 = €74,278",
                "→ This demonstrates why the max cap matters — it prevents runaway liability for long-tenured employees",
            ],
        },
    ],
    "notes": [
        "We model objective dismissal (20 days/yr, capped at 360 days) as the standard severance path. Unfair dismissal is rare and handled case by case.",
        "Pagas extra (extraordinary payments) are 2 extra monthly payments per year, typically paid in June and December. They are included in the severance salary basis.",
        "Prorated pagas extra must also be paid out at termination (final settlement / finiquito).",
        "15-day notice period required for objective dismissals. No notice required for disciplinary dismissals.",
    ],
}

ws_es = wb.create_sheet("ES_Indefinite")
build_country_sheet(ws_es, "Spain", es_data)


# ========== TEMPLATE TAB ==========
template_data = {
    "country": "",
    "country_hint": "Enter country name and code, e.g., 'Germany (DE)'",
    "contract_type": "",
    "entity_type": None,
    "mta_default": "No",
    "mta_days": "",
    "mta_tiers": [
        ("", "", "", ""),
    ],
    "mta_max_cap": "",
    "components": [
        {
            "label": "Redundancy Pay",
            "is_default": True,
            "tiers": [
                ("", "", "", ""),
                ("", "", "", ""),
            ],
            "max_cap": "",
        }
    ],
    "vacation": {
        "enabled": "",
        "minimum": "",
        "tiers": [
            ("", "", "", ""),
        ],
        "max_cap": "",
    },
    "comp_vars": [
        ("Gross Base Salary", "", "", ""),
        ("", "", "", ""),
        ("", "", "", ""),
        ("", "", "", ""),
        ("", "", "", ""),
    ],
    "notes": [
        "",
        "",
        "",
    ],
}

ws_tmpl = wb.create_sheet("Template")
ws_tmpl.sheet_properties.tabColor = "D97706"
build_country_sheet(ws_tmpl, "[Country Name]", template_data)

output_path = "/Users/sanjana/Desktop/termination-liability-config/Termination_Liability_Config.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
